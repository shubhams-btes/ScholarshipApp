from multiprocessing.dummy import connection
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.forms import modelformset_factory
from .forms import QuestionForm, CollegeForm, ExamScheduleForm, CollegeOfficialForm,CollegeOfficialEditForm
from tests.models import Result, Question
from admin_panel.models import College, CollegeOfficial, ExamSchedule, ExamScheduleHistory
from students.models import Student
from tests.models import Result, Question, ExamProgress
from utils.qr_utils import generate_qr_attachment
import datetime
from django.utils.timezone import make_aware, get_default_timezone
from django.contrib.auth import logout as auth_logout
import json
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import user_passes_test
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font
from django.views.decorators.cache import never_cache
from django.utils.dateparse import parse_date
from django.utils import timezone
from datetime import datetime
import os
from django.core.paginator import Paginator
from django.db.models import Q, Prefetch
from collections import defaultdict
from django.template.loader import render_to_string
from django.core.mail import EmailMultiAlternatives,get_connection
from django.db.models import Count
from django.utils import timezone
from django.conf import settings
import logging
import threading

logger = logging.getLogger(__name__)
# -----------------------------
# Decorators
# -----------------------------
def superuser_required(view_func):
    return user_passes_test(lambda u: u.is_superuser)(view_func)

# -----------------------------
# Login
# -----------------------------
def login(request):
    from django.contrib.auth import authenticate, login as auth_login
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_superuser:
            auth_login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, "Invalid credentials or not a superuser.")
    return render(request, 'admin_panel/login.html')


def logout(request):
    """
    Logs out the current user and redirects to the admin login page.
    """
    auth_logout(request)
    return redirect('admin_login')  # Replace 'admin_login' with your login view URL name

# -----------------------------
# Dashboard
# -----------------------------
@never_cache
@superuser_required
def dashboard(request):
    college_query = request.GET.get('college')
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    schedules = ExamScheduleHistory.objects.all()

    if college_query:
        schedules = schedules.filter(college__name__icontains=college_query)

    if from_date_str:
        from_date = parse_date(from_date_str)
        if from_date:
            from_dt = timezone.make_aware(datetime.combine(from_date, datetime.min.time()))
            schedules = schedules.filter(quiz_date__gte=from_dt)

    if to_date_str:
        to_date = parse_date(to_date_str)
        if to_date:
            to_dt = timezone.make_aware(datetime.combine(to_date, datetime.max.time()))
            schedules = schedules.filter(quiz_date__lte=to_dt)

    schedules = schedules.select_related('college').order_by('-quiz_date')
    
     # Pagination
    paginator = Paginator(schedules, 5)  # 10 rows per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'admin_panel/dashboard.html', {
        'page_obj': page_obj,
        'college_query': college_query,
        'from_date': from_date_str,
        'to_date': to_date_str,
        'title': 'Dashboard'
    })


# -----------------------------
# College Management
# -----------------------------

@superuser_required
def college_management(request):
    q = request.GET.get("q", "").strip().lower()
    colleges = College.objects.prefetch_related("officials").order_by("name")

    # Filter colleges based on search
    filtered_colleges = []
    for college in colleges:
        # College matches?
        college_matches = q in college.name.lower() if q else True

        # Filter officials
        matching_officials = []
        for off in college.officials.all():
            if (
                (q in off.name.lower() if off.name else False) or
                (q in off.email.lower() if off.email else False) or
                college_matches
            ):
                matching_officials.append(off)

        if matching_officials or college_matches:
            # Add a dummy official if no real officials
            if not matching_officials:
                class DummyOfficial:
                    def __init__(self, college):
                        self.college = college
                        self.name = None
                        self.email = None
                        self.is_active = None
                        self.id = None
                matching_officials.append(DummyOfficial(college))

            # Attach matching officials to college for template
            college.matching_officials = matching_officials
            filtered_colleges.append(college)

    # Pagination on colleges
    paginator = Paginator(filtered_colleges, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "admin_panel/college_management.html", {
        "colleges": page_obj,  # template loops over colleges
        "q": q
    })


@superuser_required
def add_college(request):
    if request.method == "POST":
        college_form = CollegeForm(request.POST, prefix="college")
        official_form = CollegeOfficialForm(request.POST, prefix="official")

        if college_form.is_valid() and official_form.is_valid():
            college = college_form.save()
            official = official_form.save(commit=False)
            official.college = college
            official.save()
            messages.success(request, "College and primary official added successfully.")
            return redirect('college_management')
    else:
        college_form = CollegeForm(prefix="college")
        official_form = CollegeOfficialForm(prefix="official")

    return render(
        request,
        "admin_panel/add_edit_college.html",
        {
            "college_form": college_form,
            "official_form": official_form,
            "title": "Add College",
        },
    )


@superuser_required
def add_official(request, college_id=None):
    if college_id:
        college = get_object_or_404(College, pk=college_id)
    else:
        college = None

    if request.method == "POST":
        form = CollegeOfficialForm(request.POST)
        if form.is_valid():
            official = form.save(commit=False)
            if college:
                official.college = college
                official.is_active = True  # New officials are active by default
            official.save()
            messages.success(request, "Official added successfully.")
            return redirect("college_management")
    else:
        form = CollegeOfficialForm(initial={"college": college} if college else None)

    return render(request, "admin_panel/add_edit_official.html", {"form": form, "title": "Add College Official"})

@superuser_required
def edit_official(request, pk):
    official = get_object_or_404(CollegeOfficial, pk=pk)
    if request.method == 'POST':
        form = CollegeOfficialEditForm(request.POST, instance=official)
        if form.is_valid():
            form.save()
            messages.success(request, "Official updated successfully.")
            return redirect('college_management')
    else:
        form = CollegeOfficialEditForm(instance=official)

    return render(request, 'admin_panel/add_edit_official.html', {
        'form': form,
        'title': f'Edit Official: {official.name}'
    })


@superuser_required
@require_POST
def toggle_college_official(request, pk):
    official = get_object_or_404(CollegeOfficial, pk=pk)
    official.is_active = not official.is_active
    official.save()
    status = "activated" if official.is_active else "deactivated"
    messages.success(request, f"{official.name} ({official.college.name}) has been {status}.")
    return redirect('college_management')

@superuser_required
def exam_schedule_management(request):
    now = timezone.now()
    q = request.GET.get("q", "").strip()

    colleges = (
        College.objects
        .select_related('exam_schedule', 'exam_schedule__current_event')   # OneToOne + its live event, one JOIN
        .order_by('name')
    )
    if q:
        colleges = colleges.filter(name__icontains=q)

    rows = []
    for college in colleges:
        # OneToOne: access raises DoesNotExist if the college has no schedule yet
        schedule = getattr(college, 'exam_schedule', None)
        rows.append({"college": college, "schedule": schedule})

    # active (has a live current event) first, then by that event's date, then name
    max_date = timezone.make_aware(datetime.max)
    rows.sort(
        key=lambda r: (
            not (r["schedule"] and r["schedule"].is_active),                       # active events first
            r["schedule"].quiz_date if (r["schedule"] and r["schedule"].quiz_date) else max_date,
            r["college"].name.lower(),
        )
    )

    paginator = Paginator(rows, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'admin_panel/quiz_management.html', {
        "rows": page_obj,
        "page_obj": page_obj,
        "q": q,
        "now": timezone.localtime(now),
    })

@superuser_required
def add_exam_schedule(request):
    college_id = request.POST.get('college') or request.GET.get('college_id')
    college = get_object_or_404(College, pk=college_id)

    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('quiz_management')

    quiz_date_str = request.POST.get('quiz_date')
    if not quiz_date_str:
        messages.error(request, "Quiz date is required.")
        return redirect('quiz_management')

    try:
        naive_dt = datetime.strptime(quiz_date_str, "%Y-%m-%dT%H:%M")
        aware_dt = make_aware(naive_dt, get_default_timezone())
    except Exception:
        messages.error(request, "Invalid date/time format.")
        return redirect('quiz_management')

    if aware_dt < timezone.now():
        messages.error(request, "You cannot select a past date for the quiz.")
        return redirect('quiz_management')

    # The stable pointer row (one per college).
    schedule, _ = ExamSchedule.objects.get_or_create(college=college)

    event = schedule.current_event

    if event and event.is_active:
        # UPDATE the live occurrence in place (fix the time of the current event)
        event.quiz_date = aware_dt
        event.registration_enabled = True
        event.quiz_enabled = False
        event.save(update_fields=['quiz_date', 'registration_enabled', 'quiz_enabled'])
        messages.success(request, f"Exam schedule updated for {college.name}.")
    else:
        # No live event yet → create one and point the schedule at it.
        # Make sure no stale active occurrence lingers for this college.
        ExamScheduleHistory.objects.filter(college=college, is_active=True).update(is_active=False)
        event = ExamScheduleHistory.objects.create(
            college=college,
            quiz_date=aware_dt,
            is_active=True,
            registration_enabled=True,
            quiz_enabled=False,
        )
        schedule.current_event = event
        schedule.save(update_fields=['current_event'])
        messages.success(request, f"Exam schedule created for {college.name}.")

    return redirect('quiz_management')


def finalize_abandoned_attempts(event):
    """Finalize students who started THIS event's quiz but never submitted.
    `event` is the ExamScheduleHistory occurrence being disabled."""
    from tests.models import ExamProgress
    now = timezone.now()
    count = 0

    # Only students bound to THIS occurrence — never the college's other events.
    stale = ExamProgress.objects.filter(
        student__exam_schedule=event,
        end_time__lt=now,
    ).select_related('student')

    for progress in stale:
        student = progress.student

        # Already has a result for this occurrence → just clean up progress.
        if Result.objects.filter(student=student, exam_schedule=event).exists():
            progress.delete()
            continue

        answers = progress.answers or {}
        score = 0
        for qid_str, value in answers.items():
            if not qid_str.isdigit():
                continue
            qid = int(qid_str)
            if qid not in progress.question_ids:
                continue
            try:
                q = Question.objects.get(id=qid)
            except Question.DoesNotExist:
                continue
            if str(value) == str(q.correct_option):
                score += 1

        Result.objects.create(
            student=student,
            exam_schedule=event,
            quiz_date=event.quiz_date,          # stamp from the occurrence
            score=score,
            total_questions=len(progress.question_ids),   # int, not str
        )
        progress.delete()
        count += 1

    return count


@superuser_required
@require_POST
def toggle_quiz_status(request, pk):
    schedule = get_object_or_404(ExamSchedule, pk=pk)
    event = schedule.current_event

    if not event:
        messages.error(request, f"No active event for {schedule.college.name}.")
        return redirect('quiz_management')

    if event.quiz_enabled:
        # DISABLE: flip the occurrence fully inactive. Do NOT touch quiz_date.
        event.quiz_enabled = False
        event.registration_enabled = False
        event.is_active = False
        event.save(update_fields=['quiz_enabled', 'registration_enabled', 'is_active'])

        finalized = finalize_abandoned_attempts(event)   # resolve by occurrence, not (college, date)
        schedule.current_event = None
        schedule.save(update_fields=['current_event'])
        messages.success(
            request,
            f"Quiz disabled for {schedule.college.name}. "
            f"{finalized} abandoned attempt(s) finalized."
        )
    else:
        # ENABLE: turn the quiz on for the live occurrence.
        event.quiz_enabled = True
        event.save(update_fields=['quiz_enabled'])
        messages.success(request, f"Quiz enabled for {schedule.college.name}.")

    return redirect('quiz_management')

# -----------------------------
# Quiz Management (optional simplified)
# -----------------------------
@superuser_required
@require_POST
def share_registration_link(request, schedule_id):

    schedule = get_object_or_404(ExamSchedule,pk=schedule_id)

    link = (
        f"{settings.SITE_URL}/register/"
        f"?schedule_id={schedule.id}"
    )

    schedule.registration_link = link
    schedule.save(update_fields=["registration_link"])
    registration_qr_cid, registration_qr = generate_qr_attachment(link,filename="registration_qr.png")
    emails = list(
        schedule.college.officials
        .filter(is_active=True)
        .values_list("email", flat=True)
    )

    
    
    if not emails:
        messages.error(
            request,
            "No active college officials found."
        )
        return redirect("quiz_management")
    
    

    event = schedule.current_event
    
    if not event.registration_enabled:
        messages.error(request, f"Registration is closed for {schedule.college.name}.")
        return redirect("quiz_management")
    
    
    context = {
        "college_name": schedule.college.name,
        "registration_link": link,
        "registration_qr_cid": registration_qr_cid,
        "quiz_date": schedule.quiz_date.strftime("%d-%m-%Y %I:%M %p"),
        "site_name": settings.SITE_NAME
    }

    html_content = render_to_string(
        "emails/register_email.html",
        context
    )

    email = EmailMultiAlternatives(
        subject=(
            f"{settings.SITE_NAME} Registration - "
            f"{schedule.college.name}"
        ),
        body="Please view this email in HTML format.",
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=emails,
    )

    email.attach_alternative(
        html_content,
        "text/html"
    )
    email.attach(registration_qr)
    try:
        email.send()

        messages.success(
            request,
            f"Registration link sent to {schedule.college.name} officials."
        )

    except Exception as e:
        messages.error(
            request,
            f"Failed to send email: {e}"
        )

    return redirect("quiz_management")

def _send_quiz_links(college_name, students_data, link, quiz_date_str):
    connection = get_connection()
    connection.open()
    sent = 0
    try:
        for s in students_data:
            try:
                context = {
                    "student_name": s["name"],
                    "college_name": college_name,
                    "hall_ticket": s["hall_ticket"],
                    "quiz_link": link,
                    "quiz_date": quiz_date_str,
                    "access_time": "10 minutes before the BTES TalentQuest",
                    "site_name": settings.SITE_NAME,
                }
                html_content = render_to_string("emails/quiz_link.html", context)
                email = EmailMultiAlternatives(
                    subject=f"Your Quiz Link & Hall Ticket - {settings.SITE_NAME}",
                    body="Please view this email in HTML format.",
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[s["email"]],
                    connection=connection,
                )
                email.attach_alternative(html_content, "text/html")
                email.send()
                sent += 1
            except Exception:
                logger.exception("Quiz link failed for %s", s["email"])
    finally:
        connection.close()
    logger.info("Quiz links: sent %s/%s for %s", sent, len(students_data), college_name)


@superuser_required
@require_POST
def share_quiz_link(request, schedule_id):
    schedule = get_object_or_404(ExamSchedule, pk=schedule_id)

    event = schedule.current_event
    if not event:
        messages.error(request, f"No active event for {schedule.college.name}. Set a schedule first.")
        return redirect("quiz_management")

    link = f"{settings.SITE_URL}/login/?schedule_id={schedule.id}"   # stable id — correct
    schedule.quiz_link = link
    schedule.save(update_fields=["quiz_link"])

    local_quiz_time = timezone.localtime(event.quiz_date)
    if not event.is_active:
        messages.error(request, f"This event is disabled for {schedule.college.name}.")
        return redirect("quiz_management")

    # Students bound to THIS occurrence — the live event only.
    students = Student.objects.filter(exam_schedule=event)
    if not students.exists():
        messages.error(request, "No registered students found.")
        return redirect("quiz_management")

    students_data = [
        {"name": s.name, "email": s.email, "hall_ticket": s.hall_ticket}
        for s in students
    ]

    threading.Thread(
        target=_send_quiz_links,
        args=(
            schedule.college.name,
            students_data,
            link,
            local_quiz_time.strftime("%d-%m-%Y %I:%M %p"),
        ),
        daemon=True,
    ).start()

    messages.success(
        request,
        f"Sending quiz links to {len(students_data)} students. "
        "They will arrive over the next couple of minutes."
    )
    return redirect("quiz_management")

@superuser_required
@require_POST
def update_quiz_date(request, pk):
    schedule = get_object_or_404(ExamSchedule, pk=pk)

    event = schedule.current_event
    if not event:
        messages.error(request, f"No active event for {schedule.college.name}. Create a schedule first.")
        return redirect('quiz_management')

    quiz_date_str = request.POST.get('quiz_date')
    if not quiz_date_str:
        messages.error(request, "Quiz date is required.")
        return redirect('quiz_management')

    try:
        naive_dt = datetime.strptime(quiz_date_str, "%Y-%m-%dT%H:%M")
        aware_dt = make_aware(naive_dt, get_default_timezone())
    except Exception:
        messages.error(request, "Invalid date/time format.")
        return redirect('quiz_management')

    if aware_dt < timezone.now():
        messages.error(request, "You cannot select a past date for the quiz.")
        return redirect('quiz_management')

    event.quiz_date = aware_dt
    event.save(update_fields=['quiz_date'])
    messages.success(request, f"Quiz date updated for {schedule.college.name}.")
    return redirect('quiz_management')

@superuser_required
@require_POST
def toggle_registration(request, pk):
    schedule = get_object_or_404(ExamSchedule, pk=pk)

    event = schedule.current_event
    if not event:
        messages.error(request, f"No active event for {schedule.college.name}.")
        return redirect('quiz_management')

    event.registration_enabled = not event.registration_enabled
    event.save(update_fields=['registration_enabled'])

    status = "opened" if event.registration_enabled else "closed"
    messages.success(request, f"Registration {status} for {schedule.college.name}.")
    return redirect('quiz_management')

# -----------------------------
# Results per College
# -----------------------------
@superuser_required
def college_results(request, schedule_id):
    schedule = get_object_or_404(ExamScheduleHistory, pk=schedule_id)
    college = schedule.college

    # Get results only for this exam schedule
    results = Result.objects.filter(exam_schedule=schedule).order_by("-score")

    # Apply optional filters
    cutoff = request.GET.get("cutoff")
    top_n = request.GET.get("top_n")

    filtered_results = results
    if cutoff:
        filtered_results = filtered_results.filter(score__gte=int(cutoff))
    if top_n:
        filtered_results = filtered_results[:int(top_n)]

    # Pagination: 10 results per page
    paginator = Paginator(filtered_results, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "admin_panel/results.html", {
        "college": college,
        "schedule": schedule,
        "page_obj": page_obj,
        "cutoff": cutoff,
        "top_n": top_n,
    })


@superuser_required
def college_registrations(request, schedule_id):
    # Get the quiz schedule
    schedule = get_object_or_404(ExamScheduleHistory, pk=schedule_id)
    college = schedule.college

    # Fetch all students registered for this quiz schedule
    registered_students = Student.objects.filter(exam_schedule=schedule).order_by('name')

    # Pagination: 10 students per page
    paginator = Paginator(registered_students, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "admin_panel/registrations.html", {
        "college": college,
        "schedule": schedule,
        "page_obj": page_obj
    })

# -----------------------------
# Question Management
# -----------------------------
@superuser_required
def manage_questions(request):
    questions = Question.objects.all().order_by('-id')
    paginator = Paginator(questions, 10)  # paginate by 10
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'admin_panel/manage_questions.html', {'page_obj': page_obj})


@superuser_required
def add_question(request):
    if request.method == 'POST':
        form = QuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)

            question.uploaded_by = request.user

            question.save()
            messages.success(request, "Question added successfully.")
            return redirect('manage_questions')
    else:
        form = QuestionForm()
    return render(request, 'admin_panel/add_edit_question.html', {'form': form, 'title': 'Add Question'})


@superuser_required
def edit_question(request, pk):
    question = get_object_or_404(Question, pk=pk)
    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)
        if form.is_valid():

            question = form.save(commit=False)

            question.updated_by = request.user

            question.save()

            messages.success(request, "Question updated successfully.")

            return redirect('manage_questions')
    else:
        form = QuestionForm(instance=question)
    return render(request, 'admin_panel/add_edit_question.html', {'form': form, 'title': 'Edit Question'})


@superuser_required
@require_POST
def toggle_question(request, pk):
    question = get_object_or_404(Question, pk=pk)
    question.is_active = not question.is_active
    question.save(update_fields=['is_active'])
    status = "enabled" if question.is_active else "disabled"
    messages.success(request, f"Question #{question.id} has been {status}.")
    return redirect('manage_questions')

@superuser_required
@require_POST
def toggle_all_questions(request, action):
    if action == 'enable':
        Question.objects.update(is_active=True)
        messages.success(request, "All questions have been enabled.")
    elif action == 'disable':
        Question.objects.update(is_active=False)
        messages.success(request, "All questions have been disabled.")
    else:
        messages.error(request, "Invalid action.")
    return redirect('manage_questions')

@superuser_required
@require_http_methods(["GET", "POST"])
def upload_questions(request):
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            messages.error(request, "Please select a file to upload.")
            return redirect('upload_questions')

        # --- FIX 1: Validate allowed extensions ---
        allowed_extensions = ['.json', '.txt']
        file_ext = os.path.splitext(file.name)[1].lower()

        if file_ext not in allowed_extensions:
            messages.error(request, "Invalid file type. Only .json or .txt files are allowed.")
            return redirect('upload_questions')

        try:
            MAX_UPLOAD_BYTES = 2 * 1024 * 1024  # 2 MB
            if file.size > MAX_UPLOAD_BYTES:
                messages.error(request, "File too large. Maximum allowed size is 2 MB.")
                return redirect('upload_questions')
            # Read file
            data = file.read().decode('utf-8')

            # --- FIX 2: Validate JSON format ---
            try:
                questions_data = json.loads(data)
            except json.JSONDecodeError:
                messages.error(request, "Invalid JSON format. Ensure the file contains valid JSON.")
                return redirect('upload_questions')

            if not isinstance(questions_data, list):
                messages.error(request, "Invalid JSON format: root must be a list of questions.")
                return redirect('upload_questions')

            allowed_categories = ["TECHNICAL", "REASONING"]

            count = 0
            for q in questions_data:
                if not isinstance(q, dict):
                    continue

                # Extract fields safely
                category = q.get("category", "").strip().upper()
                question_text = q.get("question_text", "").strip()
                option_1 = q.get("option_1", "").strip()
                option_2 = q.get("option_2", "").strip()
                option_3 = q.get("option_3", "").strip()
                option_4 = q.get("option_4", "").strip()
                correct_option = q.get("correct_option")

                # --- NEW: Validate category ---
                if category not in allowed_categories:
                    messages.error(
                        request,
                        f"Invalid category '{category}'. Allowed categories: TECHNICAL or REASONING only."
                    )
                    return redirect('upload_questions')

                # Validate basic question fields
                if (
                    not question_text or 
                    correct_option not in [1, 2, 3, 4] or
                    not all([option_1, option_2, option_3, option_4])
                ):
                    continue

                # Create question
                Question.objects.create(

                    category=category,

                    question_text=question_text,

                    option_1=option_1,

                    option_2=option_2,

                    option_3=option_3,

                    option_4=option_4,

                    correct_option=correct_option,

                    uploaded_by=request.user

                )

                count += 1

            messages.success(request, f"{count} questions uploaded successfully!")
            return redirect('manage_questions')

        except Exception as e:
            # messages.error(request, f"Error processing file: {str(e)}")
            return redirect('upload_questions')

    # GET request
    return render(request, 'admin_panel/upload_questions.html')




# Export Routes

@superuser_required
def export_registrations(request, schedule_id):
    schedule = get_object_or_404(ExamScheduleHistory, pk=schedule_id)
    students = Student.objects.filter(exam_schedule=schedule).order_by('name')

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Registrations_{schedule.college.name}"

    # Header row
    headers = ["ID", "Name", "Email", "College", "Contact Number"]
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index, value=header.upper())
        cell.font = Font(bold=True)

    # Data rows
    for idx, student in enumerate(students, start=1):
        row = [
            idx,
            student.name.upper() if student.name else "",
            student.email.upper() if student.email else "",
            student.exam_schedule.college.name.upper() if student.exam_schedule.college.name else "",
            student.mobile_number.upper() if student.mobile_number else ""
        ]
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=idx+1, column=col_index, value=value)

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Registrations_{schedule.college.name}_{schedule.quiz_date.date()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@superuser_required
def export_results(request, schedule_id):
    schedule = get_object_or_404(ExamScheduleHistory, pk=schedule_id)
    results = Result.objects.filter(exam_schedule=schedule).order_by('-score')

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Results_{schedule.college.name}"

    # Header row
    headers = ["ID", "Student Name", "Email", "Score","College", "Contact Number"]
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index, value=header.upper())
        cell.font = Font(bold=True)

    # Data rows
    for idx, result in enumerate(results, start=1):
        row = [
            idx,
            result.student.name.upper() if result.student.name else "",
            result.student.email.upper() if result.student.email else "",
            result.score,
            result.exam_schedule.college.name.upper() if result.exam_schedule.college.name else "",
            result.student.mobile_number.upper() if result.student.mobile_number else ""
        ]
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=idx+1, column=col_index, value=value)

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Results_{schedule.college.name}_{schedule.quiz_date.date()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@superuser_required
@require_POST
def reset_student_session(request, student_id):
    student = get_object_or_404(Student, pk=student_id)

    # Clear ONLY the login lock. Do NOT touch ExamProgress — that holds the
    # student's saved answers, question set, and deadline, so they resume
    # exactly where they left off after logging back in.
    student.current_session = None
    student.save(update_fields=["current_session"])

    messages.success(
        request,
        f"Login reset for {student.name} ({student.hall_ticket}). "
        f"They can now log in again."
    )

    # Redirect back to the same registrations page.
    schedule_id = request.POST.get("schedule_id")
    if schedule_id:
        return redirect("college_registrations", schedule_id=schedule_id)
    return redirect("dashboard")
